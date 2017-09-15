from openpyxl import Workbook
from openpyxl import load_workbook
import json



def response_dict(response):
	data_dict = {}
	data = json.loads(response.read())

	# want to create test data	
	# with open('vlb_json_test.json', 'w') as json_dump:
	# 	json_dump.write(str(data))

	data_dict['content'] = data

	if response.code == 200:
		data_dict['code'] = response.code		
	else:
		data_dict['code'] = response.code		

	# # test data	
	# with open('C:\\Code\\trade_cat_scripts\\tests\\dataset\\vlb_test_pickle_error.txt', 'wb') as test_pickle:
	# 	pickle.dump(data_dict, test_pickle)

	return data_dict

def get_sheetdata(file):

	wb = load_workbook(file)
	names = wb.get_sheet_names()
	return wb[names[0]] # data returned

def xlreport(title):

	buk = Workbook()
	outsheet = buk.active
	outsheet.title = title
	return buk, title

def create_headers(outsheet, trader):

	if trader == 'VLB':

		outsheet.cell(row=1, column=1, value="BFLUX ISBN")
		outsheet.cell(row=1, column=2, value="BFLUX Title")
		outsheet.cell(row=1, column=3, value="BFLUX Imprint")
		outsheet.cell(row=1, column=4, value=trader+" ISBN")
		outsheet.cell(row=1, column=5, value=trader+" Title")
		outsheet.cell(row=1, column=6, value="ISBN Match")
		outsheet.cell(row=1, column=7, value="BFLUX Promo Status")
		outsheet.cell(row=1, column=8, value="BFLUX Del Status")
		outsheet.cell(row=1, column=9, value=trader+" Availability")
		outsheet.cell(row=1, column=10, value=trader+" Avail Desc.")
		outsheet.cell(row=1, column=11, value="price_eur_br")
		outsheet.cell(row=1, column=12, value=trader+" Price DE")
		outsheet.cell(row=1, column=13, value="Price DE Match")
		outsheet.cell(row=1, column=14, value="price_eur_br_a")
		outsheet.cell(row=1, column=15, value=trader+" Price AT")
		outsheet.cell(row=1, column=16, value="Price AT Match")
		outsheet.cell(row=1, column=17, value="price_chf")
		outsheet.cell(row=1, column=18, value=trader+" Price CH")
		outsheet.cell(row=1, column=19, value="Price CH Match")
		outsheet.cell(row=1, column=20, value="# of "+trader+" Prices")
		outsheet.cell(row=1, column=21, value=trader+" Front Cover")
		outsheet.cell(row=1, column=22, value="BFLUX cover status")

	if trader == 'AmznDE':

		outsheet.cell(row=1, column=1, value="BFLUX ISBN")
		outsheet.cell(row=1, column=2, value="BFLUX Title")
		outsheet.cell(row=1, column=3, value="BFLUX Imprint")
		outsheet.cell(row=1, column=4, value=trader+" ISBN")
		outsheet.cell(row=1, column=5, value=trader+" Title")
		outsheet.cell(row=1, column=6, value="ISBN Match")
		outsheet.cell(row=1, column=7, value="BFLUX Promo Status")
		outsheet.cell(row=1, column=8, value="BFLUX Del Status")
		outsheet.cell(row=1, column=9, value=trader+" Availability")
		outsheet.cell(row=1, column=10, value="price_eur_br")
		outsheet.cell(row=1, column=11, value=trader+" Price DE")
		outsheet.cell(row=1, column=12, value=trader+" Front Cover")
		outsheet.cell(row=1, column=13, value="BFLUX cover status")

	return outsheet

def add_row(trader, outsheet, row_n, item, book):

	if trader == 'VLB':
		if item.isbn == book.isbn:
			match = True
		else:
			match = False		

		outsheet.cell(row=row_n, column=1, value=item.isbn)
		outsheet.cell(row=row_n, column=2, value=item.title)
		outsheet.cell(row=row_n, column=3, value=item.imprint)	
		outsheet.cell(row=row_n, column=4, value=book.isbn)
		outsheet.cell(row=row_n, column=5, value=book.title)
		outsheet.cell(row=row_n, column=6, value=match)
		outsheet.cell(row=row_n, column=7, value=item.promo_status)
		outsheet.cell(row=row_n, column=8, value=item.del_status)
		outsheet.cell(row=row_n, column=9, value=book.availability)
		outsheet.cell(row=row_n, column=10, value=book.availability_desc)
		outsheet.cell(row=row_n, column=11, value=item.price_DE)	
		outsheet.cell(row=row_n, column=12, value=book.price_DE)	

		if item.price_DE == book.price_DE:
			match = True
		else:
			match = False

		outsheet.cell(row=row_n, column=13, value=match)
		outsheet.cell(row=row_n, column=14, value=item.price_AT)
		outsheet.cell(row=row_n, column=15, value=book.price_AT)

		if item.price_AT == book.price_AT:
			match = True
		else:
			match = False

		outsheet.cell(row=row_n, column=16, value=match)
		outsheet.cell(row=row_n, column=17, value=item.price_CH)
		outsheet.cell(row=row_n, column=18, value=book.price_CH)

		if item.price_CH == book.price_CH:
			match = True
		else:
			match = False

		outsheet.cell(row=row_n, column=19, value=match)
		outsheet.cell(row=row_n, column=20, value=book.prices)
		outsheet.cell(row=row_n, column=21, value=book.front_cover)
		outsheet.cell(row=row_n, column=22, value=item.cover_status)

	if trader == 'AmznDE':
		print(trader)
		if item.isbn == book.isbn:
			match = True
		else:
			match = False

		outsheet.cell(row=row_n, column=1, value=item.isbn)
		outsheet.cell(row=row_n, column=2, value=item.title)
		outsheet.cell(row=row_n, column=3, value=item.imprint)
		outsheet.cell(row=row_n, column=4, value=book.isbn)
		outsheet.cell(row=row_n, column=5, value=book.title)
		outsheet.cell(row=row_n, column=6, value=match)
		outsheet.cell(row=row_n, column=7, value=item.promo_status)
		outsheet.cell(row=row_n, column=8, value=item.del_status)
		outsheet.cell(row=row_n, column=9, value=book.availability)
		outsheet.cell(row=row_n, column=10, value=item.price_DE)
		outsheet.cell(row=row_n, column=11, value=book.price)
		outsheet.cell(row=row_n, column=12, value=book.cover)
		outsheet.cell(row=row_n, column=13, value=item.cover_status)
	
	return outsheet

def file_splitter(row_count):
	row_count_each_file = 12000
	if row_count > row_count_each_file:		
		remain = row_count%row_count_each_file
		filler = row_count_each_file - remain
		return row_count + filler
	else:
		return row_count

def get_worksheets(file):
	worksheets = []
	workbook = load_workbook(file)
	for worksheet in workbook.worksheets:
		if worksheet.title != 'DQL':
			worksheets.append(worksheet)
	return worksheets

# print(file_splitter(11000))

	

