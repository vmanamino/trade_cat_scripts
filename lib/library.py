from openpyxl import Workbook
from openpyxl import load_workbook

def get_sheetdata(file):

	wb = load_workbook(file)
	names = wb.get_sheet_names()
	return wb[names[0]] # data returned

def xlreport(title):

	buk = Workbook()
	outsheet = buk.active
	outsheet.title = title
	return outsheet

def create_headers(outsheet, trader):

	outsheet.cell(row=1, column=1, value="BFLUX ISBN")
	outsheet.cell(row=1, column=2, value="BFLUX Title")
	outsheet.cell(row=1, column=3, value=trader+" ISBN")
	outsheet.cell(row=1, column=4, value=trader+" Title")
	outsheet.cell(row=1, column=5, value="ISBN Match")
	outsheet.cell(row=1, column=6, value="BFLUX Promo Status")
	outsheet.cell(row=1, column=7, value=trader+" Availability")
	outsheet.cell(row=1, column=8, value="price_eur_br")
	outsheet.cell(row=1, column=9, value="Price DE")
	outsheet.cell(row=1, column=10, value="Price Match")
	outsheet.cell(row=1, column=11, value="# of "+trader+" Prices")

	return outsheet