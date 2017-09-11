from openpyxl import load_workbook
import string
import json, os
import sys
# append path to keys for import
sys.path.append('C:\\Code\\trade_cat_scripts')
import keys
import urllib.request
import datetime
# import pickle

def get_product_data(isbn):

	url = 'http://api.vlb.de/api/v1/product/'+str(isbn)+'/isbn13'
	return response_dict(get_request(url))

def add_headers(request):
	
	metadata_token = os.environ['VLB_TOKEN_METADATA']
	request.add_header('Authorization', 'Bearer '+metadata_token)
	request.add_header('Content-Type', 'application/json')
	request.add_header('Accept', 'application/json')
	return request

def get_request(url):

    request = add_headers(urllib.request.Request(url))
    try:
    	return urllib.request.urlopen(request)
    except urllib.request.HTTPError as err:
    	return err    	
	
def response_dict(response):

	# want to create test data	
	# with open('vlb_json_test.json', 'w') as json_dump:
	# 	json_dump.write(str(data))

	data_dict = {}

	try:
		data = json.loads(response.read())

	except ValueError:
		data_dict['code'] = 'response format buggy'
		data_dict['content'] = 'response format buggy'
	
	else:
		data_dict['content'] = data

		if response.code == 200:
			data_dict['code'] = response.code		
		else:
			data_dict['code'] = response.code		

	# # test data	
	# with open('C:\\Code\\trade_cat_scripts\\tests\\dataset\\vlb_test_pickle_error.txt', 'wb') as test_pickle:
	# 	pickle.dump(data_dict, test_pickle)
	finally:
		return data_dict

def get_frontcover(mediafiles):
	# get the first 04 (front cover) OR 06 (front cover High Quality)
	flag = False
	for file in mediafiles:
		if file['type'] == '06':
			flag = True
			return file['link']
		elif file['type'] == '04':
			flag = True
			return file['link']	
	if not flag:
			return "No front cover"

# add AT and CH
def dach_prices(prices):	
	today = datetime.date.today()
	flag_de = False
	flag_de_valid = False
	flag_at = False
	flag_at_valid = False
	flag_ch = False
	flag_ch_valid = False
	dachs = {'DE':'', 'AT':'', 'CH':''}

	for price in prices:		
		valid_price_date = False		
		validFrom_str = price['validFrom']
		validTo_str = price['validUntil']
		validFrom = ''
		validTo = ''
		
		if validFrom_str:
			day, month, year =  map(int, validFrom_str.split('.'))
			validFrom = datetime.date(year, month, day)	

			if validTo_str:			
				day, month, year =  map(int, validTo_str.split('.'))
				validTo = datetime.date(year, month, day)		
			
			if validFrom < today:
				if validTo:				
					if validTo > today:
						valid_price_date = True
						
				else:				
					valid_price_date = True
					

		elif validTo_str == None:		
			valid_price_date = True		

		else:			
			day, month, year =  map(int, validTo_str.split('.'))
			validTo = datetime.date(year, month, day)
			if validTo > today:			
				valid_price_date = True

		# if the price has a valid date, then assign it to a country		
		if valid_price_date:

			if price['country'] == 'DE':
				flag_de = True				
				dachs['DE'] = price['value']
			if price['country'] == 'AT':
				flag_at = True
				dachs['AT'] = price['value']
			if price['country'] == 'CH':
				flag_ch = True
				dachs['CH'] = price['value']
		# if a price does not have a valid date, then
		# if a price has not been assigned already for a specific country
		elif price['country'] == 'DE':			
			if not dachs['DE']:
				flag_de = True
				dachs['DE'] = 'No valid DE price'
		elif price['country'] == 'AT':			
			if not dachs['AT']:
				flag_de = True
				dachs['AT'] = 'No valid AT price' 
		elif price['country'] == 'CH':			
			if not dachs['CH']:
				flag_de = True
				dachs['CH'] = 'No valid CH price'		
	if not flag_de:
		dachs['DE'] = "No DE Price"
	if not flag_at:
		dachs['AT'] = "No AT Price"
	if not flag_ch:
		dachs['CH'] = "No CH Price"

	return dachs

def avail_code_desc(avail_code):
	desc = ''
	if avail_code == 'AB':
		desc = 'Cancelled'	
	elif avail_code == 'AD':
		desc = 'Available direct from publisher only'
	elif avail_code == 'CS':
		desc = 'Availability uncertain'
	elif avail_code == 'EX':
		desc = 'No longer stocked by us'
	elif avail_code == 'IP':
		desc = 'Available'
	elif avail_code == 'MD':
		desc = 'Manufactured on demand'
	elif avail_code == 'NP':
		desc = 'Not yet published'
	elif avail_code == 'NY':
		desc = 'Newly catalogued, not yet in stock'
	elif avail_code == 'OF':
		desc = 'Other format available'
	elif avail_code == 'OI':
		desc = 'Out of stock indefinitely'
	elif avail_code == 'OP':
		desc = 'Out of print'
	elif avail_code == 'OR':
		desc = 'Replaced by new edition'
	elif avail_code == 'PP':
		desc = 'Publication postponed indefinitely'	
	elif avail_code == 'RF':
		desc = 'Refer to another supplier'
	elif avail_code == 'RM':
		desc = 'Remaindered'
	elif avail_code == 'RP':
		desc = 'Reprinting'
	elif avail_code == 'RU':
		desc = 'Reprinting, undated'
	elif avail_code == 'TO':
		desc = 'Special order'
	elif avail_code == 'TP':
		desc = 'Temporarily out of stock because publisher cannot supply'
	elif avail_code == 'TU':
		desc = 'Temporarily unavailable'
	elif avail_code == 'UR':
		desc = 'Unavailable, awaiting reissue'
	elif avail_code == 'WR':
		desc = 'Will be remaindered as of (date)'
	elif avail_code == 'WS':
		desc = 'Withdrawn from sale'
	else:
		desc = 'No valid code'

	return desc

'''
function to get each sheet from Delilah excel file, in cases 
where results are greater than 65000 titles
'''
def get_worksheets(file):
	worksheets = []
	workbook = load_workbook(file)
	for worksheet in workbook.worksheets:
		if worksheet.title != 'DQL':
			worksheets.append(worksheet)
	return worksheets


'''
function to get the data from Delilah excel file
'''
def get_sheetdata(file):
	wb = load_workbook(file)
	names = wb.get_sheet_names()
	return wb[names[0]] # data returned

# print(get_product(9781430261063))
# print(get_product_data(9781484213933)['code'])
# prices = get_product_data(9781484213933)['content']['prices']
# print(dach_prices(prices)) 
# prices = get_product_data(9783658147747)['content']['prices'] # good test
# print(dach_prices(prices))
# # get_request('http://api.vlb.de/api/v1/product/9783476043306/isbn13')
# print(avail_code_desc('MD'))
# prices = get_product_data(9788132236566)['content']['prices']
# prices = get_product_data(9783658128319)['content']['prices']
# print(prices)
# print(dach_prices(prices))
# with open('C:\\Code\\trade_cat_scripts\\tests\\dataset\\vlb_dach_prickes_pickle.txt', 'wb') as test_pickle:
		# pickle.dump(prices, test_pickle)