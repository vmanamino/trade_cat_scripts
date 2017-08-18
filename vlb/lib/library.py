from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
import json, os
import sys
# append path to keys for import
sys.path.append('C:\\Code\\trade_cat_scripts')
import keys
import urllib.request
import pickle

metadata_token = os.environ['VLB_TOKEN_METADATA']

def get_product(isbn):

	url = 'http://api.vlb.de/api/v1/product/'+str(isbn)+'/isbn13'
	return response_dict(get_request(url))	

def add_headers(request):
	
	request.add_header('Authorization', 'Bearer '+metadata_token)
	request.add_header('Content-Type', 'application/json')
	request.add_header('Accept', 'application/json')
	return request

def get_request(url):

    request = add_headers(urllib.request.Request(url))
    try:
    	return urllib.request.urlopen(request)
    except urllib.request.HTTPError as err:
    	return err    	
	
def response_dict(response):
	data_dict = {}
	data = json.loads(response.read())

	# want to create test data	
	# with open('vlb_json_test.json', 'w') as json_dump:
	# 	json_dump.write(str(data))

	data_dict['content'] = data

	if response.code == 200:
		data_dict['code'] = response.code		
	else:
		data_dict['code'] = response.code		

	# # test data	
	# with open('C:\\Code\\trade_cat_scripts\\tests\\dataset\\vlb_test_pickle_error.txt', 'wb') as test_pickle:
	# 	pickle.dump(data_dict, test_pickle)

	return data_dict

def dach_prices(prices):
	flag = False
	for price in prices:
		if price['country'] == 'DE':
			flag = True
			return price['value']
	if not flag:
		return 'No German Price'

'''
function to get the data from Delilah file
'''
def get_sheetdata(file):

	wb = load_workbook(file)

	names = wb.get_sheet_names()

	return wb[names[0]] # data returned
	
# print(get_product(9783476043306))
# print(get_product(9781430261063))
# print(get_product(9781484213933))
# # get_request('http://api.vlb.de/api/v1/product/9783476043306/isbn13')


