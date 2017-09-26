from openpyxl import Workbook
import sys
sys.path.append('C:\\Code\\trade_cat_scripts\\amzn\\lib')
from library_amzn import collate_amzn_data as collate
from amzn_product import AmznProduct as Product

sys.path.append('C:\\Code\\trade_cat_scripts\\lib')
from library_common import get_sheetdata, get_worksheets
from bflux_item import BFLUXItem
from report import Report
import time
startTime = time.time()
# item_id = '9783476043306'
# data = collate(item_id)
# print(data)
# product = Product(data)
# print(product.isbn)

# report = Report('isbn check', 'AmznDE')


medium = sys.argv[1]
promotion = sys.argv[2]
year = sys.argv[3]
option = sys.argv[4]

# responseGroups = {'title_info': 'ItemAttributes', 'cover_info': 'Images', 'sales_info': 'OfferFull'}

responseGroups = {'title_info': 'ItemAttributes'}

filename = medium + '_' + promotion + '_' + year + '.xlsx'
print(filename)

log_count = 0

log = open('results\\simple_log.txt', 'w')
log.write('%s\t%s\t%s\t%s\n' % ('ISBN', 'Response Group', 'Log time', 'Log date'))

buk = Workbook()

if option == 'workbook':
	print(option)

	item_count = 1 # for the header
	worksheets = get_worksheets('dataset\\'+filename)

	for data in worksheets:
		count = data.max_row
		# range is to, not including the upper limit
		count = count + 1
		for n in range(2, count):
			item_count += 1
			log_count += 1
			log_date = time.strftime("%d:%m:%y")
			log_time = time.strftime("%I:%M:%S")
			print(log_count, end=': ')
			print(log_time)		
			item = BFLUXItem(data, n)
			product_data = collate(item.isbn)			
			log.write('%s\t%s\t%s\t%s\n' % (item.isbn, response_group, log_time, log_date))
			book = Product(product_data)
			report.generate(item_count, item, book)
# create loop on responseGroups
# create inner loop on row count
# set row count to zero, as many times as number of responseGroups, here 3
# then loop on row count as many times

elif option == 'spreadsheet':
	

	group_count = 0
	for report in responseGroups:
		response_group = responseGroups[report]
		print(option)
		data = get_sheetdata('dataset\\'+filename)
		count = data.max_row
		print(count)
		# range is to, not including the upper limit
		count = count + 1
		if response_group == 'ItemAttributes':
			isbn_check = buk.active
			isbn_check.title = 'ISBN Check'
			isbn_check.cell(row=1, column=1, value="BFLUX ISBN")
			isbn_check.cell(row=1, column=2, value="BFLUX Title")
			isbn_check.cell(row=1, column=3, value="AMZN DE ISBN")
			isbn_check.cell(row=1, column=4, value="AMZN DE Title")
			isbn_check.cell(row=1, column=5, value="Match")
			for n in range(2, count):
				time.sleep(1)
				log_count += 1
				log_date = time.strftime("%d:%m:%y")
				log_time = time.strftime("%I:%M:%S")
				print(log_count, end=': ')
				print(log_time)		
				item = BFLUXItem(data, n)
				product_data = collate(item.isbn, response_group)		
				log.write('%s\t%s\t%s\t%s\n' % (item.isbn, response_group, log_time, log_date))
				print(product_data)
				isbn_check.cell(row=n, column=1, value=item.isbn)
				isbn_check.cell(row=n, column=2, value=item.title)
				isbn_check.cell(row=n, column=3, value=product_data['isbn'])
				isbn_check.cell(row=n, column=4, value=product_data['title_info'])
				flag = False
				if item.isbn == product_data['isbn']:
					flag = True
				isbn_check.cell(row=n, column=5, value=flag)			
		if response_group == 'OfferFull':
			sales_info = buk.create_sheet(report)
			sales_info.cell(row=1, column=1, value="BFLUX ISBN")
			sales_info.cell(row=1, column=2, value="BFLUX Title")
			sales_info.cell(row=1, column=3, value="BFLUX Price")
			sales_info.cell(row=1, column=4, value="AMZN DE Price")
			sales_info.cell(row=1, column=5, value="BFLUX Availabity")
			sales_info.cell(row=1, column=6, value="AMZN DE Availability")
			for n in range(2, 15):
				time.sleep(1)
				log_count += 1
				log_date = time.strftime("%d:%m:%y")
				log_time = time.strftime("%I:%M:%S")
				print(log_count, end=': ')
				print(log_time)		
				item = BFLUXItem(data, n)
				product_data = collate(item.isbn, response_group)		
				log.write('%s\t%s\t%s\t%s\n' % (item.isbn, response_group, log_time, log_date))				
				print(product_data)
				sales_info.cell(row=n, column=1, value=item.isbn)
				sales_info.cell(row=n, column=2, value=item.title)
				sales_info.cell(row=n, column=3, value=item.price_DE)
				sales_info.cell(row=n, column=4, value=product_data['price'])
				sales_info.cell(row=n, column=5, value=item.del_status)
				sales_info.cell(row=n, column=6, value=product_data['availability'])			

	print_date = time.strftime("%d%m%y")
	print_time = time.strftime("%I%M%S")
	report_name = 'amznDE_'+medium + '_' + promotion + '_' + year + '_report_'+print_date+'_'+print_time
	buk.save('results\\'+report_name+'.xlsx')

		
'''
need to work on calling the api from here
'''
# 	for n in range(2, 100):
# 		time.sleep(10)
# 		log_count += 1
# 		log_date = time.strftime("%d:%m:%y")
# 		log_time = time.strftime("%I:%M:%S")
# 		print(log_count, end=': ')
# 		print(log_time)		
# 		item = BFLUXItem(data, n)
# 		product_data = collate(item.isbn)		
# 		log.write('%s\t%s\t%s\t%s\n' % (item.isbn, response_group, log_time, log_date))
# 		book = Product(product_data)
# 		report.generate(n, item, book) 

# print_date = time.strftime("%d%m%y")
# print_time = time.strftime("%I%M%S")

# report_name = 'amznDE_'+medium + '_' + promotion + '_' + year + '_report_'+print_date+'_'+print_time
# report.save('results\\'+report_name)
log.close()